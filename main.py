import sys
import os
from math import pi
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from PyQt5 import QtWidgets as qtw
from PyQt5 import QtCore as qtc
from PyQt5 import QtGui as qtg
from PyQt5 import uic
from datetime import datetime
import time

MyWindow, base_class = uic.loadUiType('./gui.ui')
class MainWindow(base_class):
    def __init__(self, *args,**kwargs):
        super().__init__(*args,**kwargs)
        self.ui = MyWindow()
        self.ui.setupUi(self)
        
        self.initialize()

        
    def initialize(self):
        self.ui.new_done_tick.setHidden(True)
        self.ui.existing_done_tick.setHidden(True)
        self.timer=qtc.QTimer()
        self.timer.timeout.connect(lambda: self.ui.new_done_tick.setHidden(True))
        self.timer.timeout.connect(lambda: self.ui.existing_done_tick.setHidden(True))

        #BUTTONS
        self.ui.select_txt_btn.clicked.connect(lambda: self.load_file('Txt','txt'))
        self.ui.select_existing_excel_btn.clicked.connect(lambda: self.load_file('Excel','xlsx'))
        self.ui.write_new_btn.clicked.connect(self.write_new_button_pressed)
        self.ui.write_existing_btn.clicked.connect(self.write_existing_btn_pressed)


   

    def load_file(self, format, extenion):
        file_path = qtw.QFileDialog.getOpenFileName(self, 'Open File',filter=f"{format} (*.{extenion})")
        
        if not file_path[0]:
            return None
        file_path = file_path[0]
        selected_file_name = os.path.basename(file_path)
        if format == 'Excel':
            self.ui.selected_xlsx_lbl.setText(f'Excel выбран:\n{selected_file_name}')
            self.excel_file_path = file_path
        else:
            self.ui.selected_txt_lbl.setText(f'Txt выбран:\n{selected_file_name}')
            self.txt_file_path = file_path
            self.events = self.parse_events()
            self.ui.found_events_lbl.setText(f'Найдено {len(self.events)} событий')

            for widget in self.ui.RightMenu.findChildren((qtw.QLabel,qtw.QPushButton,qtw.QLineEdit)):
                widget.setEnabled(True)

    def write_new_button_pressed(self):
        if not hasattr(self, 'txt_file_path'):
            return None
        logs = self.compose_logs_from_events()
        
        wb = Workbook()
        ws = wb.active
        alignment = Alignment(horizontal='center')
        for row in ws.iter_rows(min_row = 1, max_row = len(logs), min_col = 3, max_col = len(logs[0])+2):
            for cell in row:
                cell.value = logs[cell.row-1][cell.column-3]
                cell.alignment = alignment
        self.set_styles(ws,len(logs[0]))
        if self.ui.file_name_lineEdit.text() == 'logs timestamp':
            
            wb.save(f'logs {datetime.now().strftime("%d.%m.%Y %H-%M-%S")}.xlsx')
        else:
            wb.save(f'{self.ui.file_name_lineEdit.text()}.xlsx')
        self.ui.new_done_tick.setHidden(False)
        self.timer.start(1000)
        
        

    def write_existing_btn_pressed(self):
        if not hasattr(self, 'excel_file_path'):
            print('skip')
            return None
        logs = self.compose_logs_from_events()
        logs.pop(0)
        wb = load_workbook(self.excel_file_path)
        sh = wb.active
        
        alignment = Alignment(horizontal='center')
        end_of_old_logs = sh.max_row
        for row in sh.iter_rows(min_row = end_of_old_logs+1, max_row = len(logs) + end_of_old_logs, min_col = 3, max_col = len(logs[0])+2):
            for cell in row:
                cell.value = logs[cell.row-1-end_of_old_logs][cell.column-3]
                cell.alignment = alignment
        wb.save(os.path.basename(self.excel_file_path))
        self.ui.existing_done_tick.setHidden(False)
        self.timer.start(1000)

    def set_styles(self,ws, length):
        bold = Font(color="00000000", bold  = True)
        
        for row in ws.iter_rows(min_row = 1, max_row = 1, min_col=3, max_col = 2+length):
            for cell in row:
                cell.font = bold
                ws.column_dimensions[get_column_letter(cell.column)].width = 12
                
    def parse_events(self):
        with open(self.txt_file_path,'r',encoding='utf-8') as file:
            text = file.read()
            events = text.split('END EVENT')
            events.pop()
        return events
        

    def compose_logs_from_events(self):
        logs = [['Date', 'Origin time', 'Lat', 'Lon', 'Depth', 'DepthMIN','DepthMAX', 'M', 'AzMajor', 'Rminor', 'Rmajor', 'S', 'N stations']]
        
        for event in self.events:
            event_log = []
            event = event.strip()
            event = event[:event.find('\nAZIMUTHAL GAP')]
            
            dating = event[event.find('PROB='):event.find('NSTAT')]
            dating = dating[dating.find('(')+1:dating.find(')')]
            fulldate = datetime.strptime(dating, "%d.%m.%Y  %H.%M:%S.%f")
            date = fulldate.strftime("%d.%m.%Y")
            time = fulldate.strftime('%H:%M:%S.%f')[:-5]
            
            lat,lon = event[event.find(')')+2:event.find('PROB')].split()[:-1]
            lat,lon = float(lat.split('=')[1]), float(lon.split('=')[1])

            depths = event[event.find('DEPTH'):]  
            depth, depthmin, depthmax = [d.split('=')[1] for d in depths.split()]
            M = ''
            azimuth = event[event.find('ELLIPSE: ')+9:event.find('DEPTH')-1]
            AzMajor, Rminor, Rmajor = [d.split('=')[1] for d in azimuth.split()]
            S = float(Rminor) * float(Rmajor) * pi
            Nstat = event[event.find('NSTAT'):event.find('NPHASES')-1].split('=')[1]
            event_log = [date,time,round(lat,3),round(lon,3), int(depth), int(depthmin), int(depthmax), M, float(AzMajor), float(Rminor), float(Rmajor), round(S,2), int(Nstat)]
            
            logs.append(event_log)
        
        return logs


if __name__ == '__main__':
    app = qtw.QApplication(sys.argv)
    mw = MainWindow()
    mw.show()
    sys.exit(app.exec_())